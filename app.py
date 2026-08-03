import streamlit as st
import openpyxl
import datetime
import re
import os

# システム環境設定（不具合防止）
os.environ["STREAMLIT_SERVER_HEADLESS"] = "true"
os.environ["STREAMLIT_BROWSER_GATHER_USAGE_STATS"] = "false"

st.set_page_config(page_title="シフト配置確認システム", layout="wide")

# ────────────────────────────────────────────────────────
# 📌 週間予定表シートを実際のセル配置から直接読み込む関数
# ────────────────────────────────────────────────────────
def read_weekly_excel_schedule(ws, calendar_data, target_year, target_month):
    from openpyxl.utils import column_index_from_string

    weekday_blocks = {
        "火": (8, "C", "D", "E", "F", "G", "H"),
        "水": (8, "I", "J", "K", "L", "M", "N"),
        "木": (8, "O", "P", "Q", "R", "S", "T"),
        "金": (8, "U", "V", "W", "X", "Y", "Z"),
        "土": (58, "C", "D", "E", "F", "G", "H"),
        "日": (58, "I", "J", "K", "L", "M", "N"),
        "月": (58, "O", "P", "Q", "R", "S", "T"),
    }
    SLOTS_PER_BLOCK = 40

    def get_fill_hex(cell):
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return None
        fg = fill.fgColor
        if getattr(fg, "type", None) == "rgb":
            rgb = fg.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                return "#" + rgb[-6:]
        return None

    weekday_template = {}
    for w_str, (base_row, c_s1, c_h1, c_s2, c_h2, c_s3, c_h3) in weekday_blocks.items():
        col_map = {
            "s1": column_index_from_string(c_s1), "h1": column_index_from_string(c_h1),
            "s2": column_index_from_string(c_s2), "h2": column_index_from_string(c_h2),
            "s3": column_index_from_string(c_s3), "h3": column_index_from_string(c_h3),
        }
        template = {}
        last_color = {"s1": None, "s2": None, "s3": None}

        for offset in range(SLOTS_PER_BLOCK):
            row = base_row + offset
            total_min = 5 * 60 + offset * 30
            hour = (total_min // 60) % 24
            minute = total_min % 60
            row_key = "row1" if minute == 0 else "row2"
            template.setdefault(hour, {"row1": {}, "row2": {}})

            for key, col_idx in col_map.items():
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                val_str = str(val).strip() if val is not None else ""

                if not val_str:
                    if key in ("s1", "s2", "s3"):
                        last_color[key] = None
                    continue

                template[hour][row_key][key] = val_str

                if key in ("s1", "s2", "s3"):
                    color = get_fill_hex(cell)
                    if not color and val_str == "〃":
                        color = last_color[key]
                    if color:
                        template[hour][row_key][f"{key}_color"] = color
                    last_color[key] = color

        weekday_template[w_str] = template

    weekday_map = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}
    for d in range(1, 32):
        try:
            this_date = datetime.date(target_year, target_month, d)
        except ValueError:
            continue
        w_str = weekday_map[this_date.weekday()]
        template = weekday_template.get(w_str)
        if not template:
            continue
        for hour, rows in template.items():
            for row_key, values in rows.items():
                for k, v in values.items():
                    calendar_data[d][hour][row_key][k] = v

# ────────────────────────────────────────────────────────
# 📌 勤務表シート（1人1行・1日1列）を読み込む関数
# ────────────────────────────────────────────────────────
def read_duty_roster_schedule(ws, calendar_data, max_days):

    def is_target_ishikawa(name):
        # 「日勤」が同じ日に複数人と重複した際に除外する対象の石川(留=瑠)さんを判定。
        # 石川(雅)・石川(靖)など、別表記の他の石川さんは対象外。
        if not name:
            return False
        name = str(name).strip()
        if "石川" not in name:
            return False
        if "瑠" in name or "留" in name:
            return True
        return name == "石川"  # 「(瑠)」等の表記がまだ付いていない月は「石川」単独をこの方とみなす

    for d_num in range(1, max_days + 1):
        target_col = 2 + (d_num - 1)
        if target_col > ws.max_column: break

        # --- その日の全スタッフの入力内容を先に集計 ---
        day_entries = []
        for r_idx in range(4, 15):
            cell_val = ws.cell(row=r_idx, column=target_col).value
            if cell_val:
                cell_str = str(cell_val).strip()
                if not cell_str or cell_str in ["┃", "│", "↓", "｜", "〃"]: continue
                staff_name = ws.cell(row=r_idx, column=1).value
                if not staff_name: continue
                staff_name = str(staff_name).strip()
                day_entries.append((staff_name, cell_str))

        # --- 日勤(ヘルパー2)に複数人が重複したら、石川(留)を除外してもう片方を採用 ---
        nichi_people = [name for (name, cell_str) in day_entries if "日" in cell_str]
        excluded_names = set()
        if len(nichi_people) > 1:
            excluded_names = {name for name in nichi_people if is_target_ishikawa(name)}

        for staff_name, cell_str in day_entries:
            if staff_name in excluded_names and "日" in cell_str:
                continue  # 日勤重複のため石川(留)を除外

            assigned_h = None
            start_hour, start_min = None, 0
            end_hour, end_min = None, 0

            is_ake = "明" in cell_str or "ｍ" in cell_str or "m" in cell_str
            is_osor = "遅" in cell_str
            is_haya = "早" in cell_str or "ｆ" in cell_str or "f" in cell_str
            is_yoru = "夜" in cell_str or "mf" in cell_str
            is_nichi = "日" in cell_str

            if is_ake:
                assigned_h = "h3"
                start_hour, start_min = 5, 0
                end_hour, end_min = 8, 0
            elif is_osor:
                assigned_h = "h3"
                start_hour, start_min = 10, 0
                end_hour, end_min = 18, 30
            elif is_haya:
                assigned_h = "h1"
                start_hour, start_min = 7, 0
                end_hour, end_min = 15, 0
            elif is_nichi:
                assigned_h = "h2"
                start_hour, start_min = 14, 0
                end_hour, end_min = 17, 0
            elif is_yoru:
                assigned_h = "h1"
                start_hour, start_min = 17, 0
                end_hour, end_min = 23, 30

            if assigned_h and start_hour is not None and end_hour is not None:
                current_time = datetime.datetime(2026, 1, 1, start_hour, start_min)
                end_time = datetime.datetime(2026, 1, 1, end_hour, end_min)
                if current_time >= end_time: continue
                time_slots = []
                while current_time <= end_time:
                    time_slots.append((current_time.hour, current_time.minute))
                    current_time += datetime.timedelta(minutes=30)
                for idx, (h, m) in enumerate(time_slots):
                    row_key = "row1" if m == 0 else "row2"
                    if idx == 0 or idx == len(time_slots) - 1:
                        calendar_data[d_num][h][row_key][assigned_h] = staff_name
                    else:
                        current_val = calendar_data[d_num][h][row_key][assigned_h]
                        if not current_val or current_val == "┃":
                            calendar_data[d_num][h][row_key][assigned_h] = "┃"

# ────────────────────────────────────────────────────────
# 🖨️ UI・デザイン用CSS
# ────────────────────────────────────────────────────────
st.markdown("""
    <style>
    div[data-testid="stMainBlockContainer"] { 
        max-width: 96% !important; 
        padding: 4rem 1.5rem 2rem 1.5rem !important;
    }
    div[data-testid="stSelectbox"] label { display: none !important; }
    div[data-testid="stSelectbox"] { margin-top: 0px !important; padding-top: 0px !important; }
    div[data-testid="stSlider"] label { display: none !important; }
    div[data-testid="stSlider"] { margin-top: 0px !important; padding-top: 5px !important; }
    .stButton > button { height: 42px !important; min-height: 42px !important; font-weight: bold !important; }
    
    .calendar-table { table-layout: fixed !important; width: 100% !important; border-collapse: collapse !important; }
    .calendar-table td { vertical-align: middle !important; padding: 4px 1px !important; }
    
    .staff-name-box { 
        display: block !important; 
        white-space: normal !important; 
        word-break: break-all !important; 
        font-size: 11px !important; 
        text-align: center; 
        line-height: 1.1 !important; 
    }
    
    @media print {
        body * { visibility: hidden !important; position: static !important; }
        .print-target, .print-target * { visibility: visible !important; }
        
        .print-target {
            position: absolute !important;
            left: 0 !important;
            top: 0 !important;
            width: 100% !important;
            margin: 0 !important;
            padding: 0 !important;
        }
        .week-page {
            page-break-after: always !important;
            break-after: page !important;
            page-break-inside: avoid !important;
            break-inside: avoid !important;
        }
        .week-page:last-child {
            page-break-after: auto !important;
            break-after: auto !important;
        }
        html, body {
            margin: 0 !important;
            padding: 0 !important;
            height: auto !important;
        }
        @page { 
            size: A3 landscape; 
            margin: 8mm !important; 
        }
        .week-print-table { border: 3px solid #000 !important; width: 100% !important; }
        .week-print-table th, .week-print-table td {
            font-size: 10.5px !important; 
            padding: 1px 1px !important;
            /* 0:30まで表示しても溢れないように高さを22px→20pxに調整 */
            height: 20px !important; 
            line-height: 1.05 !important;
            border: 1px solid #000 !important;
        }
        .week-print-table .time-col { font-size: 10px !important; font-weight: bold !important; }
        .week-print-table td.day-start,
        .week-print-table th.day-start {
            border-left: 3px solid #000 !important;
        }
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("<h2 style='text-align: center; color: #333; margin-top: 0px;'>📅 シフト配置確認システム</h2>", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    weekly_file = st.file_uploader("📅 【週間予定表】のExcelを選択", type=["xlsx", "xlsm"])
with col2:
    duty_file = st.file_uploader("📋 【勤務表】のExcelを選択", type=["xlsx", "xlsm"])

calendar_data = {}
for d in range(1, 32):
    calendar_data[d] = {}
    for h in range(0, 25):
        calendar_data[d][h] = {
            "row1": {"s1": "", "h1": "", "s2": "", "h2": "", "s3": "", "h3": ""},
            "row2": {"s1": "", "h1": "", "s2": "", "h2": "", "s3": "", "h3": ""}
        }

target_year = 2026
target_month = 6

if weekly_file is not None or duty_file is not None:
    try:
        # --- 週間予定表の読み込み ---
        if weekly_file is not None:
            wb_weekly = openpyxl.load_workbook(weekly_file, data_only=True)
            weekly_sheet_name = st.selectbox("週間予定表のシートを選択してください 👇", options=wb_weekly.sheetnames, key="weekly_sheet")
            ws_weekly = wb_weekly[weekly_sheet_name]
            
            match = re.search(r'(\d+)月|\.(\d+)', weekly_sheet_name)
            if match:
                target_month = int(match.group(1) or match.group(2))
            
            read_weekly_excel_schedule(ws_weekly, calendar_data, target_year, target_month)
            st.success(f"🎉 週間予定表「{weekly_sheet_name}」を反映しました。")

        # --- 勤務表の読み込み ---
        if duty_file is not None:
            wb_duty = openpyxl.load_workbook(duty_file, data_only=True)
            duty_sheet_name = st.selectbox("勤務表のシートを選択してください 👇", options=wb_duty.sheetnames, key="duty_sheet")
            ws_duty = wb_duty[duty_sheet_name]
            
            if weekly_file is None:
                match = re.search(r'(\d+)月|\.(\d+)', duty_sheet_name)
                if match:
                    target_month = int(match.group(1) or match.group(2))

            if target_month in [4, 6, 9, 11]: max_days = 30
            elif target_month == 2: max_days = 28
            else: max_days = 31
            
            read_duty_roster_schedule(ws_duty, calendar_data, max_days)
            st.success(f"🎉 勤務表「{duty_sheet_name}」を重ねて反映しました。")

    except Exception as e:
        st.error(f"Excelの読み込みエラー: {e}")

try: start_offset = (datetime.date(target_year, target_month, 1).weekday() + 1) % 7
except: start_offset = 0

if target_month in [4, 6, 9, 11]: max_days = 30
elif target_month == 2: max_days = 28
else: max_days = 31

def get_bg(val, h_type):
    if not val or val in ["┃", "｜", "↓", "〃"]: return "#ffffff"
    if h_type == "h3": return "#ffc0cb"
    if h_type == "h1": return "#ffff00"
    return "#ffffff"

def wrap_name(val, h_type):
    if not val: return ""
    if val in ["┃", "｜", "↓", "〃"]: return str(val)
    cleaned_val = str(val).replace("（", "(").replace("）", ")")
    return f"<span class='staff-name-box'>{cleaned_val}</span>"

def get_era_label(year, month):
    reiwa_year = year - 2018
    era_str = f"R{reiwa_year}" if reiwa_year >= 1 else str(year)
    return f"{era_str}.{month}月"

era_label = get_era_label(target_year, target_month)

if weekly_file is not None or duty_file is not None:
    # 表示中の予定を直接調整するための小さな編集パネル。
    # 印刷用の表そのものは HTML なので、ここで calendar_data を更新してから
    # 下の表を描画することで、画面表示と印刷結果を常に同じ内容にする。
    name_candidates = set()
    service_codes = set("bcefhmsw")
    for day_data in calendar_data.values():
        for hour_data in day_data.values():
            for row_data in hour_data.values():
                for cell_key in ("s1", "h1", "s2", "h2", "s3", "h3"):
                    value = str(row_data.get(cell_key, "")).strip()
                    if value and value not in {"┃", "｜", "↓", "〃"}:
                        # サービス記号（例: 山田 bc）は候補名から外す。
                        base_name = re.sub(r"\\s*[bcefhmsw]{1,2}$", "", value, flags=re.IGNORECASE).strip()
                        name_candidates.add(base_name or value)
    person_options = [""] + sorted(name_candidates) + ["〃"]
    service_colours = {"白": "#ffffff", "黄": "#ffff00", "ピンク": "#ffc0cb"}
    colour_labels = {value: label for label, value in service_colours.items()}

    def slot_ref(time_label):
        if time_label == "24:00":
            return 0, "row1"
        if time_label == "0:30":
            return 0, "row2"
        hour, minute = map(int, time_label.split(":"))
        return hour, "row1" if minute == 0 else "row2"

    time_options = [f"{hour}:{minute:02d}" for hour in range(5, 24) for minute in (0, 30)] + ["24:00", "0:30"]
    with st.expander("週間表示を編集", expanded=False):
        st.caption("サービス利用者は名前・記号（1つ目は必須）・色を選べます。ヘルパーは開始時刻の名前を変えると、その勤務枠の最後にも同じ名前を入れます。")
        pick_col1, pick_col2 = st.columns(2)
        with pick_col1:
            edit_day = st.selectbox("日付", list(range(1, max_days + 1)), key="edit_day")
        with pick_col2:
            edit_time = st.selectbox("開始時刻", time_options, key="edit_time")

        edit_hour, edit_row = slot_ref(edit_time)
        edit_cells = calendar_data[edit_day][edit_hour][edit_row]
        for service_key, helper_key, label in (("s1", "h1", "1"), ("s2", "h2", "2"), ("s3", "h3", "3")):
            service_col, helper_col = st.columns(2)
            with service_col:
                current_service = str(edit_cells.get(service_key, ""))
                current_base = re.sub(r"\\s*[bcefhmsw]{1,2}$", "", current_service, flags=re.IGNORECASE).strip()
                if current_base not in person_options:
                    person_options.append(current_base)
                selected_person = st.selectbox(f"サービス {label}：名前", person_options, index=person_options.index(current_base), key=f"{edit_day}_{edit_time}_{service_key}_name")
                code_col1, code_col2, colour_col = st.columns(3)
                current_letters = re.findall(r"[bcefhmsw]", current_service.lower())[-2:]
                with code_col1:
                    code1 = st.selectbox("記号 1", ["b", "c", "e", "f", "h", "m", "s", "w"], index=["b", "c", "e", "f", "h", "m", "s", "w"].index(current_letters[0]) if current_letters else 0, key=f"{edit_day}_{edit_time}_{service_key}_code1")
                with code_col2:
                    code2 = st.selectbox("記号 2（任意）", ["なし", "b", "c", "e", "f", "h", "m", "s", "w"], index=(["なし", "b", "c", "e", "f", "h", "m", "s", "w"].index(current_letters[1]) if len(current_letters) > 1 else 0), key=f"{edit_day}_{edit_time}_{service_key}_code2")
                with colour_col:
                    current_colour = edit_cells.get(f"{service_key}_color", "#ffffff").lower()
                    colour = st.selectbox("色", list(service_colours), index=list(service_colours).index(colour_labels.get(current_colour, "白")), key=f"{edit_day}_{edit_time}_{service_key}_colour")
                if selected_person == "〃":
                    edit_cells[service_key] = "〃"
                elif selected_person:
                    edit_cells[service_key] = f"{selected_person} {code1}{'' if code2 == 'なし' else code2}"
                else:
                    edit_cells[service_key] = ""
                edit_cells[f"{service_key}_color"] = service_colours[colour]
            with helper_col:
                current_helper = str(edit_cells.get(helper_key, ""))
                if current_helper not in person_options:
                    person_options.append(current_helper)
                selected_helper = st.selectbox(f"ヘルパー {label}：名前", person_options, index=person_options.index(current_helper), key=f"{edit_day}_{edit_time}_{helper_key}_name")
                # 固定時間の勤務枠を、開始セルから次の空欄まで追跡する。
                # 途中の「〃」は残し、開始と最終セルの表示名だけを揃える。
                slot_sequence = [slot_ref(t) for t in time_options]
                start_index = slot_sequence.index((edit_hour, edit_row))
                end_index = start_index
                old_helper = current_helper
                for index in range(start_index + 1, len(slot_sequence)):
                    next_hour, next_row = slot_sequence[index]
                    next_value = str(calendar_data[edit_day][next_hour][next_row].get(helper_key, ""))
                    if next_value in {"〃", old_helper}:
                        end_index = index
                    else:
                        break
                edit_cells[helper_key] = selected_helper
                if end_index > start_index:
                    end_hour, end_row = slot_sequence[end_index]
                    calendar_data[edit_day][end_hour][end_row][helper_key] = selected_helper

    weekdays_labels = ["日", "月", "火", "水", "木", "金", "土"]
    total_weeks = (start_offset + max_days + 6) // 7  # 月内の全日数をカバーするのに必要な週数

    st.components.v1.html(f"""
        <button onclick="parent.window.print();" style="width:100%; height:45px; background-color:#e67e22; color:white; border:none; border-radius:4px; cursor:pointer; font-size:15px; font-weight:bold;">🖨️ 全{total_weeks}週分を印刷（週ごとに自動改ページ）</button>
    """, height=50)
    st.caption(f"📄 {era_label} の全{total_weeks}週分をまとめて表示しています。印刷すると1週間＝1ページで自動的に分かれて出力されます。")
    st.write("---")

    h_sheet = []
    h_sheet.append("<div class='print-target'>")

    for week_idx in range(total_weeks):
        start_d = week_idx * 7 + 1 - start_offset

        h_sheet.append(f"<div class='week-page' id='week-page-{week_idx + 1}' style='margin-bottom:22px;'>")
        h_sheet.append(f"<div style='text-align:right; font-size:12px; font-weight:bold; color:#333; padding:0 2px 2px 0;'>{era_label} 第{week_idx + 1}週</div>")

        h_sheet.append("<table class='calendar-table week-print-table' style='width:100%; border-collapse:collapse; text-align:center; font-family:sans-serif; table-layout:fixed; border:2px solid #333;'>")
        h_sheet.append("<tr style='background-color: #f0f0f0; font-weight: bold;'>")
        h_sheet.append("<td class='time-col' style='width: 4.2%; padding: 4px 0;'>時間</td>")

        for d_o_w in range(7):
            cur_d = start_d + d_o_w
            c_color = '#ff4b4b' if d_o_w == 0 else '#1c83e1' if d_o_w == 6 else '#333333'
            if 1 <= cur_d <= max_days:
                h_sheet.append(f"<td colspan='6' class='day-start' style='color: {c_color}; font-size: 12px; width: 13.68%;'>{weekdays_labels[d_o_w]} ({cur_d}日)</td>")
            else:
                h_sheet.append(f"<td colspan='6' class='day-start' style='color: #aaa; background-color: #fafafa; width: 13.68%;'>{weekdays_labels[d_o_w]} (外)</td>")
        h_sheet.append("</tr>")

        h_sheet.append("<tr style='background-color: #f9f9f9; font-size: 9px; height: 14px;'><td style='font-weight:bold; padding:0;'>-</td>")
        for _ in range(7):
            h_sheet.append("<td class='day-start' style='width:1.8%; padding:0;'>サ1</td><td style='font-weight:bold; width:2.1%; background:#fff3cd; padding:0;'>へ1</td>")
            h_sheet.append("<td style='width:1.8%; padding:0;'>サ2</td><td style='font-weight:bold; width:2.1%; background:#fff3cd; padding:0;'>へ2</td>")
            h_sheet.append("<td style='width:1.8%; padding:0;'>サ3</td><td style='font-weight:bold; width:2.1%; background:#fff3cd; padding:0;'>へ3</td>")
        h_sheet.append("</tr>")

        r_list = []
        for hour in range(5, 24):
            r_list.append((f"{hour}:00", True))
            r_list.append((f"{hour}:30", False))

        r_list.append(("24:00", True))
        r_list.append(("0:30", False))

        for idx, (t_str, is_even) in enumerate(r_list):
            b_style = "border-bottom:1px solid #333;" if not is_even else "border-bottom:1px dashed #ccc;"
            h_sheet.append(f"<tr style='{b_style}'><td class='time-col' style='background-color:#f2f2f2;'>{t_str}</td>")

            for d_o_w in range(7):
                cur_d = start_d + d_o_w
                if 1 <= cur_d <= max_days:
                    ds = calendar_data[cur_d]

                    if t_str == "24:00":
                        rd = ds[0]["row1"]
                    elif t_str == "0:30":
                        rd = ds[0]["row2"]
                    else:
                        h_num = int(t_str.split(":")[0])
                        rd = ds[h_num]["row1" if t_str.endswith(":00") else "row2"]

                    first_cell = True
                    for sk, hk in [("s1", "h1"), ("s2", "h2"), ("s3", "h3")]:
                        bg_color_s = rd.get(f"{sk}_color", "#ffffff")
                        cls = " class='day-start'" if first_cell else ""
                        h_sheet.append(f"<td{cls} style='background-color:{bg_color_s};'>{wrap_name(rd[sk], '')}</td>")
                        h_sheet.append(f"<td style='font-weight:bold; background-color:{get_bg(rd[hk], hk)};'>{wrap_name(rd[hk], hk)}</td>")
                        first_cell = False
                else:
                    h_sheet.append("<td colspan='6' class='day-start' style='background-color: #fafafa; opacity:0.1;'></td>")
            h_sheet.append("</tr>")
        h_sheet.append("</table>")
        h_sheet.append("</div>")  # /week-page

    h_sheet.append("</div>")  # /print-target

    st.html(f"<div style='border: 1px solid #999; background-color: #ffffff; border-radius: 6px; padding: 5px;'>{''.join(h_sheet)}</div>")
else:
    st.info("💡 上部のメニューから、解析元のシフトExcelファイルをアップロードしてください。「週間予定表」と「勤務表」両方を入れると合算されて表示されます。")
